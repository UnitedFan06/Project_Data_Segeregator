import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import os

file_path = "project_data.xlsx"
main_sheet_identifier = 0 # Can be 0 for first sheet, or "Your Main Sheet Name"

actual_main_sheet_name = None

if not os.path.exists(file_path):
    print(f"Error: The file '{file_path}' was not found.")
    exit()

try:
    wb = load_workbook(file_path, read_only=True)
    if isinstance(main_sheet_identifier, int):
        if main_sheet_identifier < len(wb.sheetnames):
            actual_main_sheet_name = wb.sheetnames[main_sheet_identifier]
        else:
            print(f"Error: Main sheet index {main_sheet_identifier} out of range.")
            wb.close()
            exit()
    else:
        if main_sheet_identifier in wb.sheetnames:
            actual_main_sheet_name = main_sheet_identifier
        else:
            print(f"Error: Main sheet '{main_sheet_identifier}' not found.")
            wb.close()
            exit()
    wb.close()
except InvalidFileException:
    print(f"Error: '{file_path}' is not a valid Excel file or is corrupted.")
    exit()
except Exception as e:
    print(f"Error determining main sheet name: {e}")
    exit()

try:
    project_data = pd.read_excel(file_path, sheet_name=actual_main_sheet_name)
    print(f"Loaded main sheet '{actual_main_sheet_name}'.")
except Exception as e:
    print(f"Error reading main Excel file '{actual_main_sheet_name}': {e}.")
    exit()

unique_projects = project_data['Project'].unique()

all_sheets_data = {actual_main_sheet_name: project_data}

existing_workbook_sheets = []
if os.path.exists(file_path):
    try:
        wb = load_workbook(file_path, read_only=True)
        existing_workbook_sheets = wb.sheetnames
        wb.close()
    except InvalidFileException:
        print(f"Warning: '{file_path}' not valid for inspecting sheets.")
    except Exception as e:
        print(f"Error inspecting existing sheets: {e}")

print("Processing projects:")
for project_name in unique_projects:
    if pd.isna(project_name): # Skip if project name is NaN or None
        print("Skipping entry with NaN 'Project' name.")
        continue

    sheet_name_for_project = str(project_name) + "_Projects"
    if len(sheet_name_for_project) > 31:
        sheet_name_for_project = sheet_name_for_project[:31]

    new_project_data_df = project_data[project_data['Project'] == project_name].copy()

    df_to_write_to_sheet = new_project_data_df

    if sheet_name_for_project in existing_workbook_sheets:
        print(f"  - Appending data to '{sheet_name_for_project}'")
        try:
            existing_sheet_df = pd.read_excel(file_path, sheet_name=sheet_name_for_project)
            combined_df = pd.concat([existing_sheet_df, new_project_data_df], ignore_index=True)
            df_to_write_to_sheet = combined_df.drop_duplicates()
        except Exception as e:
            print(f"    Warning: Could not read existing sheet '{sheet_name_for_project}'. Overwriting. Error: {e}")
            df_to_write_to_sheet = new_project_data_df
    else:
        print(f"  - Creating new sheet: '{sheet_name_for_project}'")

    all_sheets_data[sheet_name_for_project] = df_to_write_to_sheet

try:
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
        for sheet_name, df_data in all_sheets_data.items():
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"\nExcel file '{file_path}' updated successfully.")
except Exception as e:
    print(f"\nError writing Excel file: {e}")
